from typing import Optional, Union, Any
import datetime
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, column_index_from_string
from decimal import Decimal
from _style_work_item import _StyleWorkItemList


def convert_issue_state(state: str) -> str:
    """
    Converts the state to the table headers.

    :param str state: the issue state
    :return: the modified state.
    :rtype: str
    """
    # the issues to convert to the New/Paused
    to_new_paused = ('New', 'Paused', 'Canceled', 'Discuss')
    # the issues to convert to the Done/Test
    to_done_test = ('Done', 'Test', 'Review')
    # the issues to convert to the Verified
    to_verified = ('Closed',)

    if state in to_new_paused:
        modified_state = 'New/Paused'
    elif state in to_done_test:
        modified_state = 'Done/Test'
    elif state in to_verified:
        modified_state = 'Verified'
    else:
        print(f"Unspecified state {state} is found.")
        modified_state = state
    return modified_state


def check_coord(coord: str) -> bool:
    """
    Verify the cell coordinate.

    :param str coord: the cell coordinates
    :return: the verification flag.
    :rtype: bool
    """
    flag = False
    try:
        coordinate_to_tuple(coord)
    except TypeError as e:
        print(f'TypeError {str(e)}. Incorrect value type.\n')
        print(f'coordinate = {coord}')
    except ValueError as e:
        print(f'ValueError {str(e)}. Incorrect value.\n')
        print(f'coordinate = {coord}')
    except OSError as e:
        print(f'OSError {e.errno}, {e.strerror}.\n')
        print(f'coordinate = {coord}')
    else:
        flag = True
    finally:
        return flag


def range_coord(start_coord: str, end_coord: str):
    """
    Convert the range string with the start and end coordinates to the tuple:\n
    (min_row, min_col, max_row, max_col)\n

    :param start_coord: the start cell coordinate, str
    :param end_coord: the end cell coordinate, str
    :return: the values for iteration of the tuple[int, int, int, int] type.
    """
    # convert start coordinate
    start_row, start_column = coordinate_to_tuple(start_coord)
    # convert end coordinate
    end_row, end_column = coordinate_to_tuple(end_coord)
    # find min and max values
    min_row = min(start_row, end_row)
    max_row = max(start_row, end_row)

    min_col = min(start_column, end_column)
    max_col = max(start_column, end_column)
    return min_row, max_row, min_col, max_col


def convert_spent_time(spent_time: Any) -> Decimal:
    """
    Convert the spent time to the specified format.

    :param spent_time: the converted value.
    :return: the decimal value.
    :rtype: Decimal
    """
    return Decimal(spent_time).normalize()


def convert_datetime_date(value: Any) -> Optional[datetime.date]:
    """
    Convert the datetime or None to the date.

    :param value: the value to convert
    :return: the date value.
    :rtype: datetime.date or None
    """
    if isinstance(value, datetime.date):
        return value
    elif isinstance(value, datetime.datetime):
        return value.date()
    else:
        return None


class ExcelProp:
    """
    Define the Excel Worksheet properties.

    Constants:
        dict_headers --- mapping states and indexes;\n
        dict_headers_short --- mapping issue states and indexes;\n
        dict_attr_column --- the dictionary of the issue attributes and columns;\n
        dict_state_style --- mapping states and styles;\n

    Class params:
        dict_pyxl_row --- the dictionary of the PyXLRow instances;\n
        dict_pyxl_work_item --- the dictionary of the PyXLWorkItem instances;\n
        dict_pyxl_merged --- the dictionary of the _PyXLMerged instances;\n
        dict_state_priority --- the dictionary of the states and priorities;\n

    Params:
        ws --- the worksheet;\n
        name --- the instance name;\n
        styles --- the _StyleWorkItemList instance;\n

    Properties:
        bottom_right --- the bottom-right cell to get the bounds;\n
        max_column --- the max column index;\n
        max_column_letter --- the max column letter;\n
        max_row --- the max row value;\n
        headers --- the header cells;\n
        headers_row --- the header row values;\n
        list_state_row --- the non-empty rows with issues for all states;\n
        get_pyxl_rows --- the PyXLRow instances from the table;\n

    Functions:
        get_column_date(date) --- get the column letter for the specified date;\n
        cell_in_range(start_coord, end_coord) --- convert the cell range to the cell generator;\n
        check_empty(item) --- verify if the cell is empty;\n
        list_state_item(state) --- get the non-empty rows with issues of the state;\n
        pyxl_row_names() --- get the PyXLRow issue names;\n
        num_pyxl_rows() --- get the issue numbers;\n
        get_work_items(row) --- get the work item instances in the row;\n
        get_merged() --- get the _PyXLMerged instances;\n
        delete_empty_rows_state(state) --- delete the empty rows for the state;\n
        delete_empty_rows() --- delete the empty rows in the workspace;\n
        get_merged_by_name(issue_name) --- get the _PyXLMerged instance by the issue name;
        get_row_by_name(issue_name) --- get the row by the issue name;
        add_work_item(issue_name, date, spent_time) --- add the work item;\n
    """
    dict_headers = {'Active': 0, 'New/Paused': 1, 'Done/Test': 2, 'Verified': 3, 'Легенда': 4}
    dict_headers_short = {'Active': 0, 'New/Paused': 1, 'Done/Test': 2, 'Verified': 3}
    dict_attr_column = {"parent": "B", "summary": "D", "deadline": "E", "commentary": "NT"}
    dict_state_style = {
        "New": "basic",
        "Active": "active",
        "Paused": "paused",
        "Done": "done",
        "Test": "test",
        "Verified": "verified_closed",
        "Closed": "verified_closed",
        "Discuss": "paused",
        "Review": "paused",
        "Canceled": "verified_closed"
    }

    dict_pyxl_row = dict()
    dict_pyxl_work_item = dict()
    dict_table_cell = dict()
    dict_state_priority: dict[str, int] = {"Active": 30, "New/Paused": 10, "Done/Test": 20, "Verified": 40}

    def __init__(self, ws: Worksheet, name: str, styles: _StyleWorkItemList):
        self.ws = ws
        self.name = name
        self.styles = styles

    def __str__(self):
        return f"Worksheet is {self.ws.title}, name is {self.name}"

    def __repr__(self):
        return f"ExcelProp(ws={self.ws}, name={self.name})"

    def __hash__(self):
        return hash((self.ws, self.name))

    def __eq__(self, other):
        if isinstance(other, ExcelProp):
            return (self.ws == other.ws) and (self.name == other.name)
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, ExcelProp):
            return (self.ws != other.ws) and (self.name != other.name)
        else:
            return NotImplemented

    def get_column_by_date(self, date: datetime.date) -> str:
        """
        Get the column letter for the specified date.

        :param datetime.date date: the date
        :return: the column_letter.
        :rtype: str
        """
        cell: Cell
        for cell in self.cell_in_range("G1", "NR1"):
            if convert_datetime_date(cell.value) == date:
                return cell.column_letter

    @property
    def bottom_right(self) -> Cell:
        """
        Get the bottom-right cell.

        :return: the bottom-right cell.
        :rtype: Cell
        """
        max_row = self.ws.max_row
        max_col = self.ws.max_column
        return self.ws.cell(max_row, max_col)

    @property
    def max_column(self) -> int:
        """
        Get the max column index.

        :return: the column index.
        :rtype: int
        """
        return self.bottom_right.column

    @property
    def max_column_letter(self) -> str:
        """
        Get the max column letter.

        :return: the column letter.
        :rtype: str
        """
        return self.bottom_right.column_letter

    @property
    def max_row(self) -> int:
        """
        Get the max row value.

        :return: the row number.
        :rtype: int
        """
        return self.bottom_right.row

    def cell_in_range(self, start_coord: str, end_coord: str):
        """
        Convert the cell range to the cell generator in the range.

        :param str start_coord: the start cell coordinate
        :param str end_coord: the end cell coordinate
        :return: the generator of cells.
        """
        if check_coord(coord=start_coord) and check_coord(coord=end_coord):
            min_row, max_row, min_col, max_col = range_coord(start_coord=start_coord, end_coord=end_coord)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    coord = f'{get_column_letter(col)}{row}'
                    yield self.ws[coord]

    @property
    def headers(self) -> list[Cell]:
        """
        Get the header cells.

        :return: the cells.
        :rtype: list[Cell]
        """
        end_coord = f"B{self.max_row}"
        return [cell for cell in self.cell_in_range("B3", end_coord) if cell.value in ExcelProp.dict_headers.keys()]

    @property
    def headers_row(self) -> list[int]:
        """
        Get the header row values.

        :return: the list of header rows.
        :rtype: list[int]
        """
        return [header.row for header in self.headers]

    def _check_empty(self, item: Union[int, str, Cell]) -> bool:
        """
        Verify if the cell is empty.

        :param item: the cell value
        :type item: int or str or Cell
        :return: the emptiness flag.
        :rtype: bool
        """
        if isinstance(item, Cell):
            return True if item.value is None else False
        elif isinstance(item, str):
            return True if self.ws[f"{item}"].value is None else False
        elif isinstance(item, int):
            cell_values = (
                self.ws[f"B{item}"].value, self.ws[f"C{item}"].value,
                self.ws[f"D{item}"].value, self.ws[f"E{item}"].value)
            if any(value is not None for value in cell_values):
                return False
            else:
                return True

    def _empty_state_item(self, state: str) -> list[int]:
        """
        Get the empty rows with issues of the state.

        :param str state: the state
        :return: the empty rows with the state.
        :rtype: list[int]
        """
        low_limit = self.headers_row[self.__index(state)] + 1
        high_limit = self.headers_row[self.__index(state) + 1] - 1
        return [row for row in range(low_limit, high_limit) if self._check_empty(row)]

    def _empty_rows(self) -> list[int]:
        """
        Get all empty rows.

        :return: the reversed empty rows.
        :rtype: list[int]
        """
        empty_rows = [self._empty_state_item(state) for state in self.dict_headers_short.keys()]
        list_empty = [item for empty_list in empty_rows for item in empty_list]
        return sorted(list_empty, reverse=True)

    def delete_empty_rows_state(self):
        """Delete all empty rows except for the pre-headers."""
        for row in self._empty_rows():
            self.ws.delete_rows(row)

    def _new_row(self, state: str) -> int:
        """
        Get the row for the new issue in the table.

        :param str state: the issue state
        :return: the row number.
        :rtype: int
        """
        return self.headers_row[self.__index(state) + 1] - 1

    def get_row_by_name(self, issue: str) -> Optional[int]:
        """
        Get the table row by the issue name.

        :param str issue: the issue name
        :return: the row number.
        :rtype: int or None
        """
        if issue in self.table_cell_names():
            return self.dict_table_cell[issue].row
        return None

    def add_work_item(self, issue_name: str, date: datetime.date, spent_time: Union[int, Decimal]):
        """
        Add the work item to the table.

        :param str issue_name: the issue name
        :param date: the work item date
        :type date: datetime.date
        :param spent_time: the work item spent time
        :type spent_time: int or Decimal
        :return: None.
        """
        column = self.get_column_by_date(date)
        row = self.get_row_by_name(issue_name)
        cell: Cell = self.ws[f"{column}{row}"]
        cell.value = spent_time

    def get_date_by_cell(self, cell: Cell) -> Optional[datetime.date]:
        """
        Define the date associated with the cell.

        :param cell: the table cell
        :type cell: Cell
        :return: the date.
        :rtype: datetime.date or None
        """
        column = cell.column_letter
        raw_date = self.ws[f"{column}1"].value
        return convert_datetime_date(raw_date)

    def __index(self, state: str) -> int:
        """
        Get the state index.

        :param str state: the state
        :return: the header index.
        :rtype: int
        """
        return self.dict_headers[state]

    def _cell_state_item(self, state: str) -> list[Cell]:
        """
        Get the cell for the state.

        :param str state: the state
        :return: the cells.
        :rtype: list[Cell]
        """
        start_row = self.dict_headers[self.__index(state)] + 1
        end_row = self.dict_headers[self.__index(state) + 1]
        start_coord = f"C{start_row}"
        end_coord = f"C{end_row}"
        return [cell for cell in self.cell_in_range(start_coord, end_coord)]

    def _cell_states(self) -> list[Cell]:
        """
        Get the cell for the TableCell instances.

        :return: the cells.
        :rtype: list[Cell]
        """
        cell_states = []
        for state in self.dict_headers_short.keys():
            start_row = self.dict_headers[self.__index(state)] + 1
            end_row = self.dict_headers[self.__index(state) + 1]
            start_coord = f"C{start_row}"
            end_coord = f"C{end_row}"
            cell: Cell
            for cell in self.cell_in_range(start_coord, end_coord):
                cell_states.append(cell)
        return cell_states

    def table_cells(self):
        """
        Get the TableCell instances.

        :return: the TableCell instances.
        :rtype: list[TableCell]
        """
        return [TableCell(self, cell) for cell in self._cell_states()]

    def table_cell_names(self) -> list[str]:
        """
        Get the issue names.

        :return: the issue names.
        :rtype: list[str]
        """
        return [table_cell.issue for table_cell in self.table_cells()]


class TableCell:
    def __init__(self, excel_prop: ExcelProp, cell: Cell):
        self.excel_prop = excel_prop
        self.cell = cell
        self.issue = str(cell.value)

        self.excel_prop.dict_table_cell[self.issue] = self

    def __str__(self):
        return f"Table cell: {self.cell.coordinate}"

    def __repr__(self):
        return f"TableCell({self.excel_prop}, {self.cell})"

    def __hash__(self):
        return hash((self.excel_prop.name, self.cell.coordinate, self.issue))

    def __key(self):
        return self.excel_prop.name, self.cell.coordinate, self.issue

    def __eq__(self, other):
        if isinstance(other, TableCell):
            return self.__key() == other.__key()
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, TableCell):
            return self.__key() != other.__key()
        else:
            return NotImplemented

    def __lt__(self, other):
        if isinstance(other, TableCell):
            return self.row < other.row
        else:
            return NotImplemented

    def __gt__(self, other):
        if isinstance(other, TableCell):
            return self.row > other.row
        else:
            return NotImplemented

    def __le__(self, other):
        if isinstance(other, TableCell):
            return self.row <= other.row
        else:
            return NotImplemented

    def __ge__(self, other):
        if isinstance(other, TableCell):
            return self.row >= other.row
        else:
            return NotImplemented

    @property
    def ws(self):
        return self.excel_prop.ws

    @property
    def row(self):
        return self.cell.row

    def _get_shift_col_idx(self, column: str) -> int:
        return column_index_from_string(column) - self.cell.col_idx

    def _shift_cell(self, value: int) -> Optional[Cell]:
        col_idx = self.cell.col_idx + value
        if col_idx > 1:
            column_letter = get_column_letter(col_idx)
            return self.ws[f"{column_letter}{self.row}"]
        else:
            print(f"ValueError, the shift {value} is out of bounds.")
            return None

    @property
    def state(self):
        for state in self.excel_prop.dict_headers_short.keys():
            if self.row in self.excel_prop._cell_state_item(state):
                return state
            else:
                continue
        return None

    def _commentary(self) -> str:
        """
        Get the commentary.

        :return: the commentary.
        :rtype: str
        """
        return self._shift_cell(self._get_shift_col_idx("NT")).value

    def _summary(self) -> str:
        """
        Get the issue summary.

        :return: the summary.
        :rtype: str
        """
        return self._shift_cell(1).value

    def _deadline(self) -> Optional[datetime.date]:
        """
        Get the issue deadline if exists.

        :return: the deadline.
        :rtype: datetime.date or None
        """
        return self._shift_cell(2).value

    def _parent(self):
        """
        Get the parent issue name if exists.

        :return: the parent issue name.
        :rtype: str
        """
        return self._shift_cell(-1).value

    def pyxl_row(self) -> tuple[str, str, str, Optional[str], Optional[datetime.date]]:
        """
        Get the values to operate with the Issue instances.

        :return: the attribute values.
        :rtype: tuple[str, str, str, Optional[str], Optional[datetime.date]]
        """
        return self.issue, self.state, self._summary(), self._parent(), self._deadline()

    def _cell_range(self):
        """
        Specify the cell generator for the range.

        :return: iterating through the cells.
        """
        return self.excel_prop.cell_in_range(f"G{self.row}", f"NR{self.row}")

    def _pyxl_cells(self) -> list[Cell]:
        """
        Get the work item cells.

        :return: the cells.
        :rtype: list[Cell]
        """
        return [cell for cell in self._cell_range() if cell.value is not None]

    def pyxl_work_items(self) -> list[tuple[str, datetime.date, Decimal]]:
        """
        Get the work items.

        :return: the work items.
        :rtype: list[tuple[str, datetime.date, Decimal]]
        """
        return [self._mapping_cell_work_item(cell) for cell in self._pyxl_cells()]

    def cell_last(self) -> Cell:
        """
        Get the last work item cell.

        :return: the last work item cell.
        :rtype: Cell
        """
        max_column = max(cell.column_letter for cell in self._pyxl_cells())
        return self.ws[f"{max_column}{self.row}"]

    def work_item_last(self) -> tuple[str, datetime.date, Decimal]:
        """
        Get the last work item.

        :return: the work item.
        :rtype: tuple[str, datetime.date, Decimal]
        """
        cell: Cell = self.cell_last()
        return self._mapping_cell_work_item(cell)

    def set_cell_style(self, style_name: str, cell: Cell):
        """

        :param style_name:
        :param cell:
        :return:
        """
        self.excel_prop.styles.set_style(style_name, cell.coordinate)

    def _mapping_cell_work_item(self, cell: Cell) -> tuple[str, datetime.date, Decimal]:
        """

        :param cell:
        :return:
        """
        date = self.excel_prop.get_date_by_cell(cell)
        return self.issue, date, Decimal(cell.value).normalize()

    def _mapping_work_item_cell(self, work_item: tuple[str, datetime.date, Decimal]) -> Cell:
        """

        :param work_item:
        :return:
        """
        _, date, _ = work_item
        column = self.excel_prop.get_column_by_date(date)
        return self.ws[f"{column}{self.row}"]

    def compare_cell_work_item(self, cell: Cell, work_item: tuple[str, datetime.date, Decimal]):
        """

        :param cell:
        :param work_item:
        :return:
        """
        if cell not in self._pyxl_cells() or work_item not in self.pyxl_work_items():
            return False
        work_item_cell = self._mapping_work_item_cell(work_item)
        return cell.coordinate == work_item_cell.coordinate

    def add_work_item(self, work_item: tuple[str, datetime.date, Decimal]):
        """

        :param work_item:
        :return:
        """
        _, date, spent_time = work_item
        cell = self._mapping_work_item_cell((self.issue, date, spent_time))
        cell.data_type = "n"
        cell.value = spent_time

    def _proper_cell_style(self, cell: Cell) -> str:
        """
        Get the cell style based on the state.

        :param cell: the cell
        :type cell: Cell
        :return: the style name.
        :rtype: str
        """
        if cell not in self._pyxl_cells():
            return cell.style.name
        elif self.excel_prop.get_date_by_cell(cell) == self._deadline():
            return "deadline"
        elif cell != self.cell_last():
            return "active"
        elif self.state is not None:
            return self.excel_prop.dict_state_style[self.state]
        else:
            return "basic"

    def _proper_work_item_style(self, work_item: tuple[str, datetime.date, Decimal]) -> str:
        """
        Get the proper work item style based on the state.

        :param work_item: the work item
        :type work_item: tuple[str, datetime.date, Decimal]
        :return: the style name
        :rtype: str
        """
        _, date, spent_time = work_item
        cell = self._mapping_work_item_cell((self.issue, date, spent_time))
        return self._proper_cell_style(cell)


def main():
    pass


if __name__ == "__main__":
    main()
